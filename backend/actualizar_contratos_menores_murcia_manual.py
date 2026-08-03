# encoding: utf-8
"""
Descarga contratos menores de fuentes municipales de Murcia cuyo formato
necesita librerías que NO están en requirements.txt de producción (odfpy para
ODS, openpyxl para XLSX, xlrd para XLS legado, pdfplumber para tablas en PDF)
-- mismo patrón manual/periódico que actualizar_alcaldes.py /
actualizar_retribuciones.py: se ejecuta a mano de vez en cuando (cada
trimestre/año, cuando el ayuntamiento publique un fichero nuevo) y genera
contratos_menores_murcia_manual.json, que app.py carga al arrancar y vuelca a
la tabla compartida contratos_menors_locales (ver
_cargar_contratos_menores_murcia_manual en app.py).

Fuente Álamo de Murcia NO está aquí: su fuente es CSV (módulo estándar `csv`,
sin dependencia nueva), así que se refresca sola en el cron diario --
ver buscar_en_fuentealamo_menores en app.py.

Ninguna de las fuentes de aquí tiene URL predecible por año/trimestre
(verificado 2026-08-03: la carpeta de subida no coincide con el periodo que
describe el fichero) -- por eso todas las funciones scrapean la página de
listado en cada ejecución en vez de adivinar la URL.

Molina de Segura publica los años 2022-2024 en XLS legado (formato binario
OLE2, Composite Document Format vía Apache POI) y solo 2025 en XLSX real
-- verificado con `file` sobre los bytes descargados, no basta con mirar la
extensión de la URL (2022 tiene ADEMÁS una copia en XLSX en un subdominio
distinto, `sedeelectronica.molinadesegura.es`, pero 2023/2024 no tienen
equivalente). Por eso hace falta `xlrd` (que dejó de soportar .xlsx en la
v2.0 pero sigue leyendo .xls perfectamente) además de `openpyxl`.

Lorquí publica un único PDF acumulativo (2015-2026, no por trimestre/año) con
una tabla real de 5 columnas (OBJETO/DURACIÓN/IMPORTE ADJUDICACIÓN/
ADJUDICATARIO/FECHA ADJUDICACIÓN). Aviso importante tras la investigación de
Albudeite (que resultó tener sobre todo anuncios/pliegos SIN adjudicatario,
no decretos de adjudicación): verificado que este PDF de Lorquí SÍ es una
tabla de contratos ya adjudicados de verdad (columna ADJUDICATARIO real en
todas las filas), no anuncios previos. `pdftotext -layout` no sirve aquí
porque las celdas envuelven a varias líneas y las columnas no quedan
alineadas de forma fiable entre filas -- `pdfplumber` sí reconstruye la
tabla real fila a fila. Aun así, la detección de rejilla de pdfplumber no es
perfecta en todas las páginas: algunas filas salen con columnas de más (con
huecos `None` intercalados) -- normalizar_fila_lorqui() los limpia y dedupa;
midiendo contra el documento completo esto recupera el 98,2% de las filas
(553 de 563) a un registro limpio de 5 campos. Sin NIF ni nº de expediente
(la fuente no los publica). Nombres de personas físicas ya en orden
correcto -- no hace falta ninguna variante invertida como en Molina de Segura.

Lorca publica un PDF POR TRIMESTRE (no acumulativo como Lorquí), pero con un
problema real descubierto al investigar (2026-08-03): **el formato de la
tabla ha cambiado al menos dos veces entre 2021 y 2026**:
- 2021-2023 (y los primeros trimestres de 2024): texto corrido SIN tabla
  real detectable por pdfplumber (0 tablas por página) -- los campos
  (Tercero/CIF, Denominación Social, Importe, Concepto, dos fechas de
  registro) vienen pegados sin separador fiable. NO soportado por este
  script.
- Desde algún trimestre de 2024 en adelante (confirmado en 4T-2024, 1T/2T-2026):
  tabla real de 5 columnas `CIF, RAZONSOCIAL, OBJETO, IMPORTE, DURACION`
  -- SIN columna de fecha por contrato (a diferencia del formato antiguo,
  que sí tenía fechas). Es el único formato que procesa este script.
- Por eso `actualizar_lorca()` NO filtra por año fijo como las demás fuentes:
  comprueba la cabecera real de cada PDF (¿la primera fila de la tabla es
  literalmente ['CIF', 'RAZONSOCIAL', 'OBJETO', 'IMPORTE', 'DURACION']?) y
  se salta entero cualquier fichero que no la tenga -- así se adapta solo
  si el ayuntamiento vuelve a cambiar el formato, en vez de producir basura.
  Al no haber fecha por contrato, se usa el primer día del trimestre (según
  el propio nombre del fichero) como fecha aproximada -- una limitación real
  de la fuente, no una decisión de diseño.
- Volumen medido: solo el 1T-2026 tiene 1.383 filas (tasa de normalización
  99,9%) -- mucho más grande que cualquier otra fuente de Murcia salvo Girona.
- Águilas se investigó también (2026-08-03) y se aparcó: la página real solo
  tiene un trimestre publicado (1T-2024, ~8 filas), congelado desde entonces
  -- no compensa el esfuerzo de un parser para ese volumen.

Uso:  pip install odfpy openpyxl xlrd pdfplumber && python actualizar_contratos_menores_murcia_manual.py
"""
import hashlib
import io
import json
import re
import sys
import time

import openpyxl
import pdfplumber
import requests
import xlrd
from bs4 import BeautifulSoup
from odf.opendocument import load as odf_load
from odf.table import Table, TableRow, TableCell
from odf.text import P as odf_P
from urllib.parse import urljoin

sys.path.insert(0, __file__.rsplit("\\", 1)[0].rsplit("/", 1)[0])
from app import BASE_DIR

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/122.0.0.0 Safari/537.36",
    "Accept-Language": "es-ES,es;q=0.9",
}
DESDE_ANY = 2021
OUT_FILE = f"{BASE_DIR}/contratos_menores_murcia_manual.json"

MULA_LISTADO_URL = "https://mula.es/web/transparencia/informacion-sobre-contratos-y-convenios/"
MOLINA_LISTADO_URL = ("https://transparencia.molinadesegura.es/publicidad-activa/"
                       "informacion-sobre-contratacion-convenios-y-subvenciones/contratos-formalizados/")
LORQUI_LISTADO_URL = "https://ayuntamientodelorqui.es/perfil-contratante/contratos-mayores-menores/"
LORCA_LISTADO_URL = "https://transparencia.lorca.es/contratos-menores/"


def _listar_enlaces(url, extension):
    """Devuelve URLs absolutas de los enlaces que acaban en `extension`. Usa
    urljoin porque no todas las fuentes traen href ya absoluto -- Lorca, por
    ejemplo, enlaza con rutas relativas ('pdf/2026/...') resueltas contra la
    URL de la propia página de listado, no contra el dominio raíz."""
    r = requests.get(url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    return sorted({urljoin(url, a["href"]) for a in soup.find_all("a", href=True)
                   if a["href"].lower().endswith(extension)})


def _num_es(valor):
    """Convierte un importe en formato español ('1.020,03') a float."""
    try:
        return float(str(valor).replace("€", "").strip().replace(".", "").replace(",", "."))
    except (TypeError, ValueError):
        return 0.0


def _num_lorca(valor):
    """Lorca usa un formato de número DISTINTO al resto de fuentes: sin
    separador de miles, punto como decimal ('48278.97 €', no '48.278,97 €')
    -- detectado en producción 2026-08-04: usar _num_es() aquí borraba el
    punto decimal asumiéndolo separador de miles y convertía 48278.97 en
    4827897.0 (un "contrato menor" de 4,8 millones de euros, imposible por
    ley -- la pista de que algo iba mal). Si alguna fila trajera coma en vez
    de punto (proveedor extranjero, poco probable pero visto NIFs franceses
    en esta fuente), se interpreta como separador decimal también."""
    limpio = str(valor).replace("€", "").strip()
    if "," in limpio and "." not in limpio:
        limpio = limpio.replace(",", ".")
    try:
        return float(limpio)
    except (TypeError, ValueError):
        return 0.0


def actualizar_mula():
    """Mula: ODS acumulativo por año, columnas NÚMERO EXP/ÁREA/CIF/PROVEEDOR/
    TIPO/OBJETO/FECHA ADJUDICACIÓN/IMPORTE IVA INCL. Nombres de personas
    físicas ya en orden correcto -- sin problema de inversión.

    OJO -- la página de listado también enlaza OTROS ODS de transparencia sin
    relación (obras públicas en ejecución, concesiones, estadísticas de
    contratos), con columnas totalmente distintas -- confirmado en producción
    2026-08-03: sin este filtro, esas filas se colaban interpretadas con el
    esquema de contratos menores y producían basura (fechas de años
    disparatados, NIF vacío). Se exige 'menor' en la URL, igual que ya se
    hace para Fuente Álamo."""
    urls = [u for u in _listar_enlaces(MULA_LISTADO_URL, ".ods") if "menor" in u.lower()]
    print(f"Mula: {len(urls)} ficheros ODS de contratos menores encontrados en el listado")
    registros = {}
    for url in urls:
        try:
            r = requests.get(url, headers=HEADERS, timeout=30)
            r.raise_for_status()
        except Exception as e:
            print(f"  Mula: {url} no disponible ({type(e).__name__})")
            continue
        doc = odf_load(io.BytesIO(r.content))
        tablas = doc.spreadsheet.getElementsByType(Table)
        if not tablas:
            continue
        filas = tablas[0].getElementsByType(TableRow)
        cabecera = None
        for fila_odf in filas:
            celdas = fila_odf.getElementsByType(TableCell)
            valores = ["".join(str(p) for p in c.getElementsByType(odf_P)) for c in celdas]
            if cabecera is None:
                cabecera = [v.strip() for v in valores]
                continue
            fila = dict(zip(cabecera, valores))
            expediente = (fila.get("NÚMERO EXP") or "").strip()
            proveedor = (fila.get("PROVEEDOR") or "").strip()
            fecha_raw = (fila.get("FECHA ADJUDICACIÓN") or "").strip()
            if not expediente or not proveedor:
                continue
            m = re.match(r"(\d{2})/(\d{2})/(\d{4})", fecha_raw)
            if not m:
                continue
            dia, mes, anio = m.groups()
            anio_actual = int(time.strftime("%Y"))
            # Cordura: una fecha de adjudicación NUNCA puede ser futura (se
            # registra a posteriori) -- detectado en producción 2026-08-03: el
            # expediente 'COME/2024/0648' trae la fecha '28/11/2029', un typo
            # real de la fuente (mismo tipo de error visto en Fuente Álamo).
            # El propio formato de expediente de Mula no es fiable para sacar
            # el año de refuerzo (cambia de orden entre ficheros: unos años
            # usan 'COME/NNNN/AAAA', otros 'COME/AAAA/NNNN'), así que aquí solo
            # se aplica la cota de cordura sobre la fecha, no un año alternativo.
            if int(anio) < DESDE_ANY or int(anio) > anio_actual:
                continue
            registros[f"Mula::{expediente}"] = {
                "id":               f"Mula::{expediente}",
                "municipio":        "Mula",
                "provincia":        "murcia",
                "fuente":           "mula",
                "organisme":        "Ayuntamiento de Mula",
                "adjudicatari":     proveedor,
                "nif":              (fila.get("CIF") or "").strip(),
                "import_num":       _num_es(fila.get("IMPORTE IVA INCL") or ""),
                "data_adjudicacio": f"{anio}-{mes}-{dia}",
                "tipus_contracte":  (fila.get("TIPO") or "").strip(),
                "descripcio":       (fila.get("OBJETO") or "").strip(),
                "codi_cpv":         "",
                "exercici":         anio,
            }
    registros = list(registros.values())
    print(f"Mula: {len(registros)} contratos menores extraídos (desde {DESDE_ANY})")
    return registros


def _iter_hojas_workbook(contenido):
    """Devuelve una lista de listas de filas (tuplas de valores), una por
    hoja, funcionando igual para XLSX real (ZIP, firma 'PK') que para XLS
    legado OLE2 (firma '\\xd0\\xcf\\x11\\xe0') -- Molina de Segura publica
    ambos formatos según el año, y la extensión de la URL no es de fiar (ver
    docstring del módulo)."""
    if contenido[:2] == b"PK":
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        return [list(wb[ws].iter_rows(values_only=True)) for ws in wb.sheetnames]
    else:
        wb = xlrd.open_workbook(file_contents=contenido)
        return [[ws.row_values(i) for i in range(ws.nrows)] for ws in wb.sheets()]


def actualizar_molina_segura():
    """Molina de Segura: XLSX/XLS anual que mezcla 'Contrato Mayor' y
    'Contrato Menor' -- se filtra por la columna 'Tipo de contratación'. El
    campo 'Contratistas' viene 'NIF - NOMBRE' (varios separados por ' | ' si
    hay UTE; aquí solo se guarda el primero, igual que se hace con las UTE de
    PSCP en app.py). OJO: para personas físicas el NOMBRE viene invertido sin
    coma ('APELLIDOS NOMBRE') -- NO se corrige aquí a propósito (nunca se
    toca el texto que se muestra en pantalla); el detector de cargos públicos
    ya prueba esa variante en tiempo de render, ver
    _variantes_nombre_para_detector en app.py."""
    urls = _listar_enlaces(MOLINA_LISTADO_URL, (".xlsx", ".xls"))
    print(f"Molina de Segura: {len(urls)} ficheros XLSX/XLS encontrados en el listado")
    registros = {}
    for url in urls:
        try:
            r = requests.get(url, headers=HEADERS, timeout=30)
            r.raise_for_status()
        except Exception as e:
            print(f"  Molina de Segura: {url} no disponible ({type(e).__name__})")
            continue
        try:
            hojas = _iter_hojas_workbook(r.content)
        except Exception as e:
            print(f"  Molina de Segura: {url} no se pudo leer ({type(e).__name__})")
            continue
        for filas_hoja in hojas:
            cabecera = None
            for row in filas_hoja:
                if cabecera is None:
                    if row and len(row) > 1 and row[1] == "Entidad Contratante":
                        cabecera = row
                    continue
                if not row or len(row) < 2 or not row[1]:
                    continue
                fila = dict(zip(cabecera, row))
                if (fila.get("Tipo de contratación") or "").strip() != "Contrato Menor":
                    continue
                expediente = str(fila.get("Número de Referencia del Contrato") or "").strip()
                if not expediente:
                    continue
                fecha_raw = str(fila.get("Fecha formalización") or "").strip()
                m = re.match(r"(\d{2})-(\d{2})-(\d{4})", fecha_raw)
                if not m:
                    continue
                dia, mes, anio = m.groups()
                # Misma cota de cordura que Mula: una fecha de formalización
                # no puede ser futura.
                if int(anio) < DESDE_ANY or int(anio) > int(time.strftime("%Y")):
                    continue
                contratistas = str(fila.get("Contratistas") or "").strip()
                primero = contratistas.split(" | ")[0]
                nif, _, nombre = primero.partition(" - ")
                importe = (fila.get("Importe total ofertado (con impuestos) (en euros)")
                           or fila.get("Importe total ofertado (sin impuestos) (en euros)") or 0)
                registros[f"Molina de Segura::{expediente}"] = {
                    "id":               f"Molina de Segura::{expediente}",
                    "municipio":        "Molina de Segura",
                    "provincia":        "murcia",
                    "fuente":           "molina-segura",
                    "organisme":        (fila.get("Entidad Contratante") or "").strip(),
                    "adjudicatari":     (nombre.strip() or primero.strip()),
                    "nif":              nif.strip(),
                    "import_num":       _num_es(importe),
                    "data_adjudicacio": f"{anio}-{mes}-{dia}",
                    "tipus_contracte":  str(fila.get("Tipo de Contrato") or "").strip(),
                    "descripcio":       str(fila.get("Objeto del Contrato") or "").strip(),
                    "codi_cpv":         str(fila.get("Código CPV del objeto del contrato") or "").strip(),
                    "exercici":         anio,
                }
    registros = list(registros.values())
    print(f"Molina de Segura: {len(registros)} contratos menores extraídos (desde {DESDE_ANY})")
    return registros


def _normaliza_fila_lorqui(row):
    """Limpia una fila cruda de pdfplumber: quita celdas None/vacías y
    colapsa duplicados consecutivos (alguna fila trae el objeto repetido dos
    veces por un artefacto de la detección de rejilla de la tabla). Devuelve
    la fila limpia solo si quedan exactamente 5 campos (objeto, duración,
    importe, adjudicatario, fecha) -- si no, se descarta como no fiable."""
    vals = [(c or "").strip() for c in row if c and str(c).strip()]
    limpio = []
    for v in vals:
        if not limpio or limpio[-1] != v:
            limpio.append(v)
    return limpio if len(limpio) == 5 else None


def actualizar_lorqui():
    """Lorquí: PDF único acumulativo (2015-2026) con tabla real de contratos
    ya adjudicados -- ver docstring del módulo para el porqué de usar
    pdfplumber en vez de pdftotext y la tasa de recuperación medida (98,2%)."""
    urls = [u for u in _listar_enlaces(LORQUI_LISTADO_URL, ".pdf") if "menor" in u.lower()]
    print(f"Lorquí: {len(urls)} ficheros PDF de contratos menores encontrados en el listado")
    registros = {}
    anio_actual = int(time.strftime("%Y"))
    for url in urls:
        try:
            r = requests.get(url, headers=HEADERS, timeout=60)
            r.raise_for_status()
        except Exception as e:
            print(f"  Lorquí: {url} no disponible ({type(e).__name__})")
            continue
        with pdfplumber.open(io.BytesIO(r.content)) as pdf:
            for pagina in pdf.pages:
                for tabla in pagina.extract_tables():
                    for row in tabla:
                        if not row or not row[0] or "OBJETO" in (row[0] or ""):
                            continue
                        fila = _normaliza_fila_lorqui(row)
                        if not fila:
                            continue
                        objeto, duracion, importe, adjudicatario, fecha_raw = fila
                        m = re.match(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", fecha_raw)
                        if not m or not re.search(r"\d", importe):
                            continue
                        dia, mes, anio = m.groups()
                        if len(anio) == 2:
                            anio = f"20{anio}"
                        # Cordura: alguna celda de fecha trae dos fechas pegadas
                        # ("28/27/2025, Modif. 12/09/2225", un typo real de la
                        # fuente con mes=27 inexistente) -- se descarta si
                        # día/mes no son un calendario válido, igual que las
                        # cotas de cordura ya aplicadas a Fuente Álamo/Mula.
                        if not (anio.isdigit() and 1 <= int(mes) <= 12 and 1 <= int(dia) <= 31):
                            continue
                        if int(anio) < DESDE_ANY or int(anio) > anio_actual:
                            continue
                        # Sin nº de expediente en esta fuente -- id por hash
                        # de contenido (objeto+adjudicatario+fecha+importe),
                        # estable frente a re-ejecuciones mientras el PDF no
                        # cambie esa fila.
                        clave_hash = hashlib.md5(
                            f"{objeto}|{adjudicatario}|{fecha_raw}|{importe}".encode("utf-8")
                        ).hexdigest()[:12]
                        # Sin columna de "tipo de contrato" en esta fuente --
                        # se reutiliza esa columna para mostrar la duración
                        # (dato real que aporta la fuente), limpiando los
                        # placeholders "----"/"-----" que usa el propio PDF
                        # para "no aplica".
                        duracion_limpia = duracion.replace("\n", " ").strip()
                        if re.fullmatch(r"-{2,}", duracion_limpia):
                            duracion_limpia = ""
                        registros[f"Lorquí::{clave_hash}"] = {
                            "id":               f"Lorquí::{clave_hash}",
                            "municipio":        "Lorquí",
                            "provincia":        "murcia",
                            "fuente":           "lorqui",
                            "organisme":        "Ayuntamiento de Lorquí",
                            "adjudicatari":     adjudicatario.replace("\n", " ").strip(),
                            "nif":              "",
                            "import_num":       _num_es(importe),
                            "data_adjudicacio": f"{anio}-{mes.zfill(2)}-{dia.zfill(2)}",
                            "tipus_contracte":  duracion_limpia,
                            "descripcio":       objeto.replace("\n", " ").strip(),
                            "codi_cpv":         "",
                            "exercici":         anio,
                        }
    registros = list(registros.values())
    print(f"Lorquí: {len(registros)} contratos menores extraídos (desde {DESDE_ANY})")
    return registros


_LORCA_CABECERA = ("CIF", "RAZONSOCIAL", "OBJETO", "IMPORTE", "DURACION")

_LORCA_MESES_TRIMESTRE = {
    "primer": "01", "1": "01",
    "segundo": "04", "2": "04",
    "tercer": "07", "3": "07",
    "cuarto": "10", "4": "10",
}


def _lorca_fecha_aprox(nombre_fichero):
    """Sin fecha por contrato en el formato nuevo (ver docstring del módulo)
    -- se aproxima al primer día del trimestre, deducido del propio nombre
    del fichero ('Primer trimestre 2026', 'Cuarto_trimestre_2024',
    'Segundo-trimestre-2024', variantes con espacio/guion/guion bajo)."""
    m = re.search(r"(primer|segundo|tercer|cuarto)[\s_-]*trimestre[\s_-]*(\d{4})",
                  nombre_fichero, re.I)
    if not m:
        return "", ""
    mes = _LORCA_MESES_TRIMESTRE[m.group(1).lower()]
    anio = m.group(2)
    return f"{anio}-{mes}-01", anio


def actualizar_lorca():
    """Lorca: un PDF por trimestre (no acumulativo). Solo se procesan los
    ficheros cuya tabla tenga la cabecera nueva (CIF/RAZONSOCIAL/OBJETO/
    IMPORTE/DURACION) -- los de formato antiguo (texto corrido, sin tabla
    real) se detectan y se saltan, ver docstring del módulo."""
    urls = [u for u in _listar_enlaces(LORCA_LISTADO_URL, ".pdf") if "trimestre" in u.lower()]
    print(f"Lorca: {len(urls)} ficheros PDF de contratos menores encontrados en el listado")
    registros = {}
    saltados = 0
    for url in urls:
        fecha_aprox, anio = _lorca_fecha_aprox(url)
        if not anio or int(anio) < DESDE_ANY:
            continue
        try:
            r = requests.get(url, headers=HEADERS, timeout=60)
            r.raise_for_status()
        except Exception as e:
            print(f"  Lorca: {url} no disponible ({type(e).__name__})")
            continue
        try:
            with pdfplumber.open(io.BytesIO(r.content)) as pdf:
                primera_tabla = pdf.pages[0].extract_tables()
                cabecera_ok = bool(primera_tabla and primera_tabla[0] and tuple(
                    (c or "").strip().upper() for c in primera_tabla[0][0][0:5]
                ) == _LORCA_CABECERA)
                if not cabecera_ok:
                    saltados += 1
                    continue
                for pagina in pdf.pages:
                    for tabla in pagina.extract_tables():
                        for row in tabla:
                            if not row or not row[0] or row[0].strip().upper() == "CIF":
                                continue
                            fila = [(c or "").strip() for c in row]
                            if len(fila) != 5:
                                continue
                            cif, razonsocial, objeto, importe, duracion = fila
                            if not razonsocial or not re.search(r"\d", importe):
                                continue
                            clave_hash = hashlib.md5(
                                f"{cif}|{razonsocial}|{objeto}|{importe}|{fecha_aprox}".encode("utf-8")
                            ).hexdigest()[:12]
                            registros[f"Lorca::{clave_hash}"] = {
                                "id":               f"Lorca::{clave_hash}",
                                "municipio":        "Lorca",
                                "provincia":        "murcia",
                                "fuente":           "lorca",
                                "organisme":        "Ayuntamiento de Lorca",
                                "adjudicatari":     razonsocial.replace("\n", " ").strip(),
                                "nif":              cif,
                                "import_num":       _num_lorca(importe),
                                "data_adjudicacio": fecha_aprox,
                                "tipus_contracte":  duracion.replace("\n", " ").strip(),
                                "descripcio":       objeto.replace("\n", " ").strip(),
                                "codi_cpv":         "",
                                "exercici":         anio,
                            }
        except Exception as e:
            print(f"  Lorca: {url} no se pudo procesar ({type(e).__name__})")
            continue
    registros = list(registros.values())
    print(f"Lorca: {len(registros)} contratos menores extraídos "
          f"({saltados} ficheros con formato antiguo descartados, desde {DESDE_ANY})")
    return registros


def main():
    todos = actualizar_mula() + actualizar_molina_segura() + actualizar_lorqui() + actualizar_lorca()
    with open(OUT_FILE, "w", encoding="utf-8") as f:
        json.dump({"generado": time.strftime("%Y-%m-%d %H:%M:%S"), "registros": todos},
                   f, ensure_ascii=False, indent=1)
    print(f"\nTotal: {len(todos)} contratos menores guardados en {OUT_FILE}")


if __name__ == "__main__":
    main()
