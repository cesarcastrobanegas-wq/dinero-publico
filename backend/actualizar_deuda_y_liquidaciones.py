# encoding: utf-8
"""
Descarga, para cada municipio de Murcia y Girona, la deuda viva municipal y
el saldo presupuestario no financiero (superávit/déficit) publicados por el
Ministerio de Hacienda, y genera backend/hacienda_eell.json.

Dos ficheros oficiales, mismo ministerio, mismo patrón de descarga directa
(XLSX público, sin sesión ni formulario) -- a diferencia de rendiciondecuentas.es,
aquí el importe SÍ está disponible: son ficheros distintos, no el mismo
portal cuyo visualizador Java ya se descartó en actualizar_cuentas_anuales.py.

1) DEUDA VIVA DE LAS ENTIDADES LOCALES (anual, foto a 31/12)
   https://www.hacienda.gob.es/es-ES/CDI/Paginas/SistemasFinanciacionDeuda/InformacionEELLs/DeudaViva.aspx
   Un único fichero "Ayuntamientos" cubre los 8.133 municipios de España
   (verificado a mano el 2026-08-04: 45/45 Murcia y 221/221 Girona presentes,
   incluidos los municipios con deuda 0 -- no son ausencias, es el valor real).

2) LIQUIDACIONES DEL PRESUPUESTO -- ESTABILIDAD PRESUPUESTARIA, SALDOS NO
   FINANCIEROS (anual, un fichero por ejercicio)
   https://www.hacienda.gob.es/es-ES/CDI/Paginas/EstabilidadPresupuestaria/InformacionCCLLs/Cumplimiento_objetivoestabilidad_EELL.aspx
   Trae, por municipio, el "Importe saldo no financiero (criterio
   presupuestario)" en euros -- el superávit/déficit real del ejercicio,
   decisión de César (2026-08-04): esta es la cifra que se muestra como
   "importe de las cuentas". No todos los municipios han remitido el
   ejercicio más reciente en el momento de la publicación (verificado con
   el fichero de 2024: 218/221 Girona, 40/45 Murcia con dato remitido) --
   por eso se prueban varios ejercicios en orden descendente y se toma el
   primero que sí tenga dato para cada municipio (misma idea que
   "último ejercicio rendido" en actualizar_cuentas_anuales.py, pero
   autocontenida en esta única fuente para no mezclar el año de un fichero
   con el importe de otro).

NINGUNA de las dos URLs de descarga es 100% predecible por patrón: el
nombre de fichero incluye la fecha de publicación, que cambia varias veces
al año a medida que llegan más remisiones -- por eso este script primero
lee las páginas índice (arriba) y extrae por regex el enlace .xlsx vigente,
en vez de construir la URL a mano (a diferencia de ISPA/actualizar_retribuciones.py,
donde sí hace falta tocar una constante a mano cada edición).

Reutiliza _emparejar_municipio() de actualizar_alcaldes.py (mismo
normalizado de acentos/apóstrofes/orden "Núcleo, Artículo" que ya resolvía
el listado del Ministerio de Política Territorial) porque el Ministerio de
Hacienda usa la misma convención de nomenclátor del INE.

Uso:  pip install openpyxl && python actualizar_deuda_y_liquidaciones.py
(openpyxl no está en requirements.txt por el mismo motivo que en los otros
scripts de este directorio: solo la usa este script, no el servidor.)
"""
import io
import json
import re
import sys
import time

import openpyxl
import requests

sys.path.insert(0, __file__.rsplit("\\", 1)[0].rsplit("/", 1)[0])
from app import BASE_DIR, normalizar
from actualizar_alcaldes import _emparejar_municipio, _sin_apostrofes_curvos, PROV_A_KEY

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/122.0.0.0 Safari/537.36",
    "Accept-Language": "es-ES,es;q=0.9",
}

BASE_HACIENDA = "https://www.hacienda.gob.es"
PAGINA_DEUDA_VIVA = (f"{BASE_HACIENDA}/es-ES/CDI/Paginas/SistemasFinanciacionDeuda/"
                     "InformacionEELLs/DeudaViva.aspx")
PAGINA_ESTABILIDAD = (f"{BASE_HACIENDA}/es-ES/CDI/Paginas/EstabilidadPresupuestaria/"
                       "InformacionCCLLs/Cumplimiento_objetivoestabilidad_EELL.aspx")

# Cuántos ejercicios probar (en orden descendente) para el saldo no
# financiero antes de dar por "sin dato" a un municipio -- 3 cubre de sobra
# el retraso de remisión observado (unos pocos municipios cada año).
EJERCICIOS_A_PROBAR = 3

OUT_FILE = f"{BASE_DIR}/hacienda_eell.json"

PROVINCIAS_CUBIERTAS = ("Murcia", "Girona")
# Nombre de provincia tal como aparece en MAYÚSCULAS en los ficheros de
# Hacienda -> nombre "Título" que usa MUNICIPIOS_POR_PROV_MIN / PROV_A_KEY.
PROVINCIA_MAYUS_A_TITULO = {"MURCIA": "Murcia", "GIRONA": "Girona"}

# Los ficheros de Hacienda escriben el artículo catalán/castellano como
# sufijo entre paréntesis ("Far d'Empordà (El)", "Torres de Cotillas
# (Las)"), mientras que el nomenclátor que ya usa esta app (MUNICIPIOS_GIRONA
# y los alias de actualizar_alcaldes.py) lo pone como sufijo con coma
# ("Far d'Empordà, el", "torres de cotillas, las") -- verificado a mano el
# 2026-08-04 al ejecutar este script por primera vez (20 municipios sin
# emparejar solo por esto). Además, el apóstrofe de nombres como "Castell
# d'Aro" llega como el carácter U+00B4 (´, ACUTE ACCENT) en vez de un
# apóstrofe recto o curvo -- ninguno de los dos está cubierto por
# _sin_apostrofes_curvos() (pensada para ’‘\`), así que se resuelve aquí.
_RE_PARENTESIS_ARTICULO = re.compile(r"^(.*)\s\((L'|El|La|Los|Las|Els|Les)\)$")


def _limpiar_nombre_hacienda(nombre):
    nombre = (nombre or "").replace("´", "'").strip()
    m = _RE_PARENTESIS_ARTICULO.match(nombre)
    if m:
        base, articulo = m.groups()
        return f"{base}, {articulo.lower()}"
    return nombre


def _descargar_xlsx(session, url):
    r = session.get(url, timeout=60)
    r.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)


def _url_absoluta(href):
    if href.startswith("http"):
        return href
    # Los espacios del path real ("sist financiacion y deuda") a veces
    # llegan ya %20-codificados en el HTML y a veces literales -- quote()
    # solo la parte de path, respetando los que ya estén codificados.
    from urllib.parse import quote
    return BASE_HACIENDA + quote(href, safe="/%")


def _descubrir_url_deuda_viva(session):
    """Busca en la página índice el primer enlace a
    'deuda-viva-ayuntamientos-*.xlsx' (o .xls) -- la página los lista con el
    año más reciente primero, verificado a mano el 2026-08-04."""
    r = session.get(PAGINA_DEUDA_VIVA, timeout=30)
    r.raise_for_status()
    m = re.search(r'href="([^"]*deuda-viva-ayuntamientos-\d+\.xlsx?)"', r.text, re.I)
    if not m:
        raise RuntimeError("no se encontró el enlace de Deuda Viva Ayuntamientos en la página índice")
    return _url_absoluta(m.group(1))


def _descubrir_urls_liquidaciones(session, n_ejercicios):
    """Busca en la página índice los N enlaces más recientes a
    'liquidacion{año}eell-estabilidadpresupuestariasaldosnofinanciero*.xlsx',
    devuelve [(ejercicio, url), ...] en orden descendente de año."""
    r = session.get(PAGINA_ESTABILIDAD, timeout=30)
    r.raise_for_status()
    encontrados = re.findall(
        r'href="([^"]*liquidacion(\d{4})eell-estabilidadpresupuestariasaldosnofinanciero[^"]*\.xlsx)"',
        r.text, re.I)
    vistos = set()
    resultado = []
    for href, anio in encontrados:
        anio = int(anio)
        if anio in vistos:
            continue
        vistos.add(anio)
        resultado.append((anio, _url_absoluta(href)))
        if len(resultado) >= n_ejercicios:
            break
    if not resultado:
        raise RuntimeError("no se encontraron enlaces de Liquidaciones/Estabilidad Presupuestaria")
    return resultado


def _procesar_deuda_viva(wb):
    """Hoja 'Datos': fila de cabecera empieza por 'Ejercicio'; columnas
    (Ejercicio, Código CCAA, CCAA, Código Provincia, Provincia, Código
    Municipio, Municipio, Deuda viva (miles de euros))."""
    ws = wb["Datos"]
    filas = ws.iter_rows(values_only=True)
    for row in filas:
        if row and row[0] == "Ejercicio":
            break
    else:
        raise RuntimeError("no se encontró la fila de cabecera 'Ejercicio' en Deuda Viva")

    resultado = {}
    sin_match = []
    for row in filas:
        if not row or not row[4]:
            continue
        provincia_mayus = str(row[4]).strip().upper()
        provincia = PROVINCIA_MAYUS_A_TITULO.get(provincia_mayus)
        if not provincia:
            continue
        nombre_crudo = _limpiar_nombre_hacienda(row[6])
        deuda_miles = row[7]
        if deuda_miles is None:
            continue
        muni = _emparejar_municipio(nombre_crudo, provincia)
        if not muni:
            sin_match.append((provincia, nombre_crudo))
            continue
        resultado[normalizar(muni)] = {
            "municipio": muni,
            "provincia": PROV_A_KEY[provincia],
            "deuda_eur": round(float(deuda_miles) * 1000, 2),
        }
    return resultado, sin_match


def _procesar_liquidaciones(wb, ejercicio):
    """Hoja 'Ayuntamientos': fila de cabecera empieza por 'Código de la
    entidad local'; columnas de interés: [4]=Nombre, [5]=Provincia,
    [7]=Importe saldo no financiero (€), [9]=Remisión de información (Sí/No)."""
    ws = wb["Ayuntamientos"]
    filas = ws.iter_rows(values_only=True)
    for row in filas:
        if row and row[0] == "Código de la entidad local":
            break
    else:
        raise RuntimeError(f"no se encontró la fila de cabecera en Liquidaciones {ejercicio}")

    resultado = {}
    sin_match = []
    for row in filas:
        if not row or not row[5]:
            continue
        provincia_mayus = str(row[5]).strip().upper()
        provincia = PROVINCIA_MAYUS_A_TITULO.get(provincia_mayus)
        if not provincia:
            continue
        if row[9] != "Si":
            continue  # no remitido todavía este ejercicio -- se prueba el anterior
        importe = row[7]
        if not isinstance(importe, (int, float)):
            continue
        nombre_crudo = _limpiar_nombre_hacienda(row[4])
        muni = _emparejar_municipio(nombre_crudo, provincia)
        if not muni:
            sin_match.append((provincia, nombre_crudo))
            continue
        resultado[normalizar(muni)] = {
            "municipio": muni,
            "provincia": PROV_A_KEY[provincia],
            "ejercicio": ejercicio,
            "importe_eur": round(float(importe), 2),
        }
    return resultado, sin_match


def main():
    session = requests.Session()
    session.headers.update(HEADERS)

    print("Localizando fichero vigente de Deuda Viva...")
    url_deuda = _descubrir_url_deuda_viva(session)
    print(f"  -> {url_deuda}")
    wb_deuda = _descargar_xlsx(session, url_deuda)
    deuda_por_muni, sin_match_deuda = _procesar_deuda_viva(wb_deuda)
    print(f"Deuda viva: {len(deuda_por_muni)} municipios emparejados"
          f"{f' (sin emparejar: {sin_match_deuda})' if sin_match_deuda else ''}")

    print(f"\nLocalizando los últimos {EJERCICIOS_A_PROBAR} ficheros de Liquidaciones/Estabilidad...")
    urls_liquidaciones = _descubrir_urls_liquidaciones(session, EJERCICIOS_A_PROBAR)
    for anio, url in urls_liquidaciones:
        print(f"  {anio} -> {url}")

    saldo_por_muni = {}
    fuente_url_por_muni = {}
    for ejercicio, url in urls_liquidaciones:
        print(f"Descargando Liquidaciones {ejercicio}...")
        wb = _descargar_xlsx(session, url)
        datos_ejercicio, sin_match_ejercicio = _procesar_liquidaciones(wb, ejercicio)
        nuevos = 0
        for clave, info in datos_ejercicio.items():
            if clave not in saldo_por_muni:  # ya cubierto por un ejercicio más reciente
                saldo_por_muni[clave] = info
                fuente_url_por_muni[clave] = url
                nuevos += 1
        print(f"  {nuevos} municipios nuevos con dato remitido en {ejercicio}"
              f" (acumulado: {len(saldo_por_muni)})"
              f"{f' -- sin emparejar: {sin_match_ejercicio}' if sin_match_ejercicio else ''}")

    for clave, info in saldo_por_muni.items():
        info["fuente_url"] = fuente_url_por_muni[clave]

    n_murcia_deuda = sum(1 for v in deuda_por_muni.values() if v["provincia"] == "murcia")
    n_girona_deuda = sum(1 for v in deuda_por_muni.values() if v["provincia"] == "girona")
    n_murcia_saldo = sum(1 for v in saldo_por_muni.values() if v["provincia"] == "murcia")
    n_girona_saldo = sum(1 for v in saldo_por_muni.values() if v["provincia"] == "girona")

    salida = {
        "generado": time.strftime("%Y-%m-%d %H:%M:%S"),
        "deuda_viva": {
            "fuente_url": url_deuda,
            "municipios": deuda_por_muni,
        },
        "saldo_no_financiero": {
            "municipios": saldo_por_muni,
        },
    }
    with open(OUT_FILE, "w", encoding="utf-8") as f:
        json.dump(salida, f, ensure_ascii=False, indent=1)

    print(f"\nDeuda viva -- Murcia: {n_murcia_deuda}/45, Girona: {n_girona_deuda}/221")
    print(f"Saldo no financiero -- Murcia: {n_murcia_saldo}/45, Girona: {n_girona_saldo}/221")
    print(f"\nGuardado en {OUT_FILE}")


if __name__ == "__main__":
    main()
